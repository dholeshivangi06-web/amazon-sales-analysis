import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------- 0. Load data ----------
# IMPORTANT: update this path if your file is not in D:Documentspython
file_path = r"D:\Documents\python\OCT DATA AMAZON 2026.xlsx"
OUTPUT_FILE = "amazon_analysis1.xlsx"
df = pd.read_excel(file_path)
data = df.copy()

# ---------- 1. Clean numeric columns and compute Total Value ----------
for c in ['item-price', 'item-tax', 'shipping-price', 'shipping-tax', 'quantity']:
    if c in data.columns:
        data[c] = pd.to_numeric(data[c], errors='coerce').fillna(0)

price_components = ['item-price', 'item-tax', 'shipping-price', 'shipping-tax']
data['Total Value'] = data[price_components].sum(axis=1) * data['quantity']

# ---------- 2. DASHBOARD sheet ----------
dashboard_cols = [
    'amazon-order-id',
    'purchase-date',
    'order-status',
    'product-name',
    'quantity',
    'ship-state',
    'ship-city',
    'item-price',
    'item-tax',
    'shipping-price',
    'shipping-tax',
    'Total Value'
]

dashboard_df = data[dashboard_cols].copy()
dashboard_df = dashboard_df.sort_values(['product-name', 'ship-state'], ascending=[True, True])

# ---------- 3. STATE_PRODUCT_SUMMARY sheet ----------
state_product_summary = (
    data.groupby(['ship-state', 'product-name'], as_index=False)
        .agg({
            'quantity': 'sum',
            'Total Value': 'sum'
        })
).sort_values('Total Value', ascending=False)

# ---------- 4. ORDER_STATUS sheet ----------
if 'item-status' in data.columns:
    status_series = data['item-status']
    label_col = 'item-status'
else:
    status_series = data['order-status']
    label_col = 'order-status'

status_counts = status_series.value_counts(dropna=False).reset_index()
status_counts.columns = [label_col, 'count']
status_counts = status_counts.sort_values('count', ascending=False)

# ---------- 5. Write ALL to ONE Excel (3 sheets) ----------
output_file = "Amazon_Analysis_Report.xlsx"
wb = Workbook()
wb.remove(wb.active)   # remove default sheet

# ===== Sheet 1: Dashboard =====
ws1 = wb.create_sheet("Dashboard", 0)

for r_idx, row in enumerate(dataframe_to_rows(dashboard_df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws1.cell(row=r_idx, column=c_idx, value=value)

        if r_idx == 1:  # header
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        elif c_idx == 1:   # amazon-order-id
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            cell.font = Font(bold=True)
        elif c_idx == 5:   # quantity
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.font = Font(bold=True)
        elif c_idx in [8, 9, 10, 11]:  # price components
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            cell.number_format = "₹#,##0.00"
        elif c_idx == 12:  # Total Value
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            cell.font = Font(bold=True)
            cell.number_format = "₹#,##0.00"

ws1.column_dimensions['A'].width = 20
ws1.column_dimensions['B'].width = 22
ws1.column_dimensions['C'].width = 25
ws1.column_dimensions['D'].width = 50
ws1.column_dimensions['E'].width = 10
ws1.column_dimensions['F'].width = 18
ws1.column_dimensions['G'].width = 20
ws1.column_dimensions['H'].width = 12
ws1.column_dimensions['I'].width = 12
ws1.column_dimensions['J'].width = 15
ws1.column_dimensions['K'].width = 15
ws1.column_dimensions['L'].width = 16
ws1.freeze_panes = "A2"

# ===== Sheet 2: State_Product_Summary =====
ws2 = wb.create_sheet("State_Product_Summary", 1)

for r_idx, row in enumerate(dataframe_to_rows(state_product_summary, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=value)

        if r_idx == 1:  # header
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        elif c_idx == 1:   # state
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.font = Font(bold=True)
        elif c_idx == 2:   # product
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        elif c_idx == 3:   # quantity
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            cell.font = Font(bold=True)
        elif c_idx == 4:   # total value
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            cell.font = Font(bold=True)
            cell.number_format = "₹#,##0.00"

ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 50
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 16
ws2.freeze_panes = "A2"


# ===== Sheet 3: Order_Status =====
ws3 = wb.create_sheet("Order_Status", 2)

for r_idx, row in enumerate(dataframe_to_rows(status_counts, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws3.cell(row=r_idx, column=c_idx, value=value)

        if r_idx == 1:  # header
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        elif c_idx == 1:  # status column
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            cell.font = Font(bold=True)
        elif c_idx == 2:  # count column
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            cell.font = Font(bold=True)

ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 15
ws3.freeze_panes = "A2"

# ---------- 6. Save Excel ----------
wb.save(output_file)

print("✅ Amazon Analysis Report generated successfully:", output_file)